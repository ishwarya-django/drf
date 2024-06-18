from django.shortcuts import render


from rest_framework.decorators import api_view,permission_classes,authentication_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from rest_framework.response import Response 
from.models import Transformer,Product,Singer,Song
from.serializers import TransformerSerializer,ProductSerializer,SingerSerializer,SongSerializer
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import SessionAuthentication,BasicAuthentication,TokenAuthentication
from django_filters.rest_framework import DjangoFilterBackend
import vonage
client = vonage.Client(key="77a4ddd4", secret="I9kjOb1EKRmLO3N4")
sms = vonage.Sms(client)
import xlwt
from django.http import HttpResponse


@api_view(['GET','POST'])
def transformer_list(request):
    """
    List all transformers, or create a new transformer
    """
    if request.method == 'GET':
        transformer = Transformer.objects.all()
       
        serializer = TransformerSerializer(transformer, many=True)
        return Response(serializer.data)
  
    elif request.method == 'POST':
        serializer = TransformerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,
                            status=status.HTTP_201_CREATED)
        return Response(serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST)




@api_view(['GET','PUT','PATCH','DELETE'])
def transformer_detail(request, pk):
    try:
        transformer = Transformer.objects.get(pk=pk)
    except Transformer.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
  
    if request.method == 'GET':
        serializer = TransformerSerializer(transformer)
        return Response(serializer.data)
  
    elif request.method == 'PUT':
        serializer = TransformerSerializer(transformer, data=request.data)
  
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'PATCH':
        serializer = TransformerSerializer(transformer,
                                           data=request.data,
                                           partial=True)
  
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST)
  
    elif request.method == 'DELETE':
        transformer.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


#   viewsets

# class Trans_former(viewsets.ModelViewSet):
#     queryset=Transformer.objects.all()
#     serializer_class= TransformerSerializer

class product(viewsets.ModelViewSet):
    queryset=Product.objects.all()
    serializer_class= ProductSerializer



class CustomAuthToken(ObtainAuthToken):

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email,
            'username':user.username,
            'firstname': user.first_name,
            'lastname': user.last_name
        })


class Singer_view(viewsets.ModelViewSet):
    queryset=Singer.objects.all()
    serializer_class= SingerSerializer

class Song_view(viewsets.ModelViewSet):
    queryset=Song.objects.all()
    serializer_class= SongSerializer


@api_view(['GET','POST'])
def add_product(request):
    if request.method == 'GET':
        transformer = Product.objects.all()
       
        serializer = ProductSerializer(transformer, many=True)
        return Response(serializer.data)
  
    elif request.method == 'POST':
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,
                            status=status.HTTP_201_CREATED)
        return Response(serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST)
    


@api_view(['GET','POST'])
def xml_upload(request):
    if request.method == "GET":
        response = HttpResponse(content_type='application/ms-excel')

        #decide file name
        response['Content-Disposition'] = 'attachment; filename="All Complaints Report.xls"'

        #creating workbook
        wb = xlwt.Workbook(encoding='utf-8')

        #adding sheet
        ws = wb.add_sheet("sheet1")

        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        # headers are bold
        font_style.font.bold = True
        #column header names, you can use your own headers here
        columns = ['Id','Name','Gender']

        #write column headers in sheet
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)
        
        data = Singer.objects.filter().order_by('-id')
        print(data)
        
        # data1 = Leads.objects.filter(lead_source='Youtube').order_by('-date')
        default_updated_staff = 'None'
        for my_row in data:
            row_num = row_num + 1
            ws.write(row_num, 0, row_num, font_style)
            ws.write(row_num, 1, my_row.name, font_style)
            ws.write(row_num, 2, my_row.gender, font_style)
        print(ws)
        wb.save(response) 
        print(wb)
        print(response)
        return response
    

from .task import test_func
from .task import send_mail_func
import json

# @api_view(['GET','POST'])
def test(request):
    # if request.method == "GET":

        test_func.delay()

        return HttpResponse("Done")


def send_mail_to_all(request):
    send_mail_func.delay()
    return HttpResponse("Sent")

from django_celery_beat.models import PeriodicTask, CrontabSchedule

def schedule_mail(request):
    schedule, created= CrontabSchedule.objects.get_or_create(hour=16, minute=48)
    print("new_updateeeeeee")
    task= PeriodicTask.objects.create(crontab=schedule,name="schedule_mail_task_"+"1",task='drf.task.send_mail_func',args=json.dumps(([2,3])))
    return HttpResponse("Done")


import pandas as pd

# @api_view(['GET','POST'])  
# def excel_data_save(request):
#     if request.method == 'POST':
#         if request.data['file']:
#             file = request.FILES['file']
#             # print(file)
#             data = pd.read_excel(file)
     
#             # print('rejin')
#             # print(data)
#             # for index1, row1 in data.iterrows():
#             #     if row1

#             for index, row in data.iterrows():
#                 print(type(row['cost']))
#                 new_cost=str(row['cost'])
#                 if new_cost == "nan":
#                     cost= 0
#                 elif type(row['cost']) is not int:  
#                     cost = 0
#                 else:
#                     cost=row['cost']

               
#                 # print(cost)
#                 if str(row['mobile_number']) == "nan":
#                     mobile_number=  None
#                 else:
#                     mobile_number = str(row['mobile_number'])
#                 Product.objects.create(product=str(row['product']),cost=cost,mobile_number=mobile_number)
#             return Response({'MESSAGE':'Datas successfully created'},status=status.HTTP_201_CREATED)  

import pandas as pd
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from io import BytesIO

@api_view(['POST'])
def excel_data_save(request):
    if request.method == 'POST':
        if request.data['file']:
            file = request.FILES['file']
            print(file)
            data = pd.read_excel(file)
            print(data)
            
            for index, row in data.iterrows():
        
                image_path = row.get('image_path')  
                print(image_path)
                
       
                if image_path:
   
                    with open(image_path, 'rb') as img_file:
                 
                        image_name = image_path.split('/')[-1] 
                        image_content = ContentFile(img_file.read())
                        uploaded_image = InMemoryUploadedFile(image_content, None, image_name, 'image/jpeg', image_content.tell(), None)
                        
                      
                        product = Transformer.objects.create(
                            name=str(row['name']),
                            # alternate_code=row['alternate_code'] if not pd.isna(row['cost']) else 0,
                            alternate_mode=row['alternate_mode'],

                            
                            image=uploaded_image  # Save image to Product model
                        )
                else:
             
                    pass
                
            return Response({'MESSAGE': 'Data successfully created'}, status=status.HTTP_201_CREATED)
        
# @api_view(['GET','POST'])
# def excel_data_view(request):

#     if request.method == "GET":
#         response = HttpResponse(content_type='application/ms-excel')
#         response['Content-Disposition'] = 'attachment; filename="All Complaints Report.xls"'
#         wb = xlwt.Workbook(encoding='utf-8')
#         ws = wb.add_sheet("sheet1")
#         row_num = 0
#         font_style = xlwt.XFStyle()
#         font_style.font.bold = True
#         columns = ['Id','Name','Altername Mode','image']
#         for col_num in range(len(columns)):
#             ws.write(row_num, col_num, columns[col_num], font_style)
#         data = Transformer.objects.filter().order_by('-id')
#         default_updated_staff = 'None'
#         for my_row in data:
#             row_num = row_num + 1
#             ws.write(row_num, 0, row_num, font_style)
#             ws.write(row_num, 1, my_row.name, font_style)
#             ws.write(row_num, 2, my_row.alternate_mode, font_style)
#             ws.write(row_num, 3, my_row.image, font_style)
#         wb.save(response) 
#         return response


# import xlwt
# from django.http import HttpResponse
# from PIL import Image as PILImage  # Import PIL's Image module as PILImage
# from io import BytesIO

# @api_view(['GET'])
# def excel_data_view(request):
#     if request.method == "GET":
#         response = HttpResponse(content_type='application/ms-excel')
#         response['Content-Disposition'] = 'attachment; filename="All_Complaints_Report.xls"'
        
#         # Create a new Workbook and sheet
#         wb = xlwt.Workbook(encoding='utf-8')
#         ws = wb.add_sheet("sheet1")

#         # Define column headers
#         columns = ['Id', 'Name', 'Alternate Mode', 'Image']
#         font_style = xlwt.XFStyle()
#         font_style.font.bold = True
        
#         # Write headers to Excel sheet
#         for col_num, col_name in enumerate(columns):
#             ws.write(0, col_num, col_name, font_style)
        
#         # Fetch data from database
#         data = Transformer.objects.all().order_by('-id').last()

#         # Initialize row counter
#         row_num = 1

#         # Loop through data and write to Excel sheet
#         for my_row in data:
#             # Write data to respective columns
#             ws.write(row_num, 0, my_row.id)
#             ws.write(row_num, 1, my_row.name)
#             ws.write(row_num, 2, my_row.alternate_mode)
            
#             # Handle image insertion using PIL
#             if my_row.image:
#                 # Example: Assuming my_row.image is a path to the image file
#                 image_path = my_row.image.path  # Adjust according to your model structure
                
#                 try:
#                     # Open the image using PIL
#                     img = PILImage.open(image_path)
                    
#                     # Resize image if necessary (optional)
#                     # img = img.resize((100, 100))  # Example resizing
                    
#                     # Create a BytesIO object to store image data
#                     img_io = BytesIO()
                    
#                     # Save image to BytesIO object
#                     img.save(img_io, format='PNG')  # Save as PNG format
                    
#                     # Add image to Excel cell as a background (xlwt workaround)
#                     ws.insert_bitmap_data(row_num, 3, img_io.getvalue())
                    
#                 except Exception as e:
#                     print(f"Error processing image: {e}")
            
#             # Increment row counter
#             row_num += 1
        
#         # Save the workbook to HttpResponse
#         wb.save(response)
#         return response



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

@api_view(['GET'])
def excel_data_view1(request):
    if request.method == "GET":
        # Create a new Workbook
        wb = Workbook()
        
        # Get the active sheet
        ws = wb.active
        ws.title = "sheet1"
        
        # Define column headers
        columns = ['Id', 'Name', 'Alternate Mode', 'Image']
        
        # Write headers to Excel sheet
        for col_num, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_num).value = col_name
        
        # Fetch data from database
        data = Transformer.objects.all().order_by('-id')
        
        # Initialize row counter
        row_num = 2  # Start from row 2 (after headers)
        
        # Loop through data and write to Excel sheet
        for my_row in data:
            # Write data to respective columns
            ws.cell(row=row_num, column=1).value = my_row.id
            ws.cell(row=row_num, column=2).value = my_row.name
            ws.cell(row=row_num, column=3).value = my_row.alternate_mode
            
            # Handle image insertion using openpyxl
            if my_row.image:
                # Example: Assuming my_row.image is a path to the image file
                image_path = my_row.image.path  # Adjust according to your model structure
                
                try:
                    # Open the image using openpyxl
                    img = XLImage(image_path)
                    
                    # Resize image if necessary (optional)
                    # img.height = 100  # Adjust height
                    # img.width = 100   # Adjust width
                    
                    # Add image to Excel sheet
                    ws.add_image(img, f'D{row_num}')  # Place image in column D, same row as data
                    
                except Exception as e:
                    print(f"Error processing image: {e}")
            
            # Increment row counter
            row_num += 1
        
        # Prepare HTTP response for downloading the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="All_Complaints_Report.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

@api_view(['GET'])
def excel_data_view2(request):
    if request.method == "GET":
        # Create a new Workbook
        wb = Workbook()
        
        # Get the active sheet
        ws = wb.active
        ws.title = "sheet1"
        
        # Define column headers
        columns = ['Id', 'Name', 'Alternate Mode', 'Image']
        
        # Write headers to Excel sheet
        for col_num, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_num).value = col_name
        
        # Fetch data from database
        data = Transformer.objects.all().order_by('-id')
        
        # Initialize row counter
        row_num = 2  # Start from row 2 (after headers)
        
        # Loop through data and write to Excel sheet
        for my_row in data:
            # Write data to respective columns
            ws.cell(row=row_num, column=1).value = my_row.id
            ws.cell(row=row_num, column=2).value = my_row.name
            ws.cell(row=row_num, column=3).value = my_row.alternate_mode
            
            # Handle image insertion using openpyxl
            if my_row.image:
                # Example: Assuming my_row.image is a path to the image file
                image_path = my_row.image.path  # Adjust according to your model structure
                
                try:
                    # Open the image using openpyxl
                    img = XLImage(image_path)
                    
                    # Calculate cell dimensions (adjust these as needed)
                    cell_width = 50
                    cell_height = 50
                    
                    # Resize image if necessary to fit within the cell
                    img.width = cell_width
                    img.height = cell_height
                    
                    # Add image to Excel sheet, fix within specific cell (e.g., D2, D3, ...)
                    ws.add_image(img, f'D{row_num}')
                    
                except Exception as e:
                    print(f"Error processing image: {e}")
            else:
                ws.cell(row=row_num, column=4).value = None

            
            # Increment row counter
            row_num += 1
        
        # Prepare HTTP response for downloading the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="All_Complaints_Report.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from io import BytesIO

@api_view(['GET'])
def excel_data_view(request):
    if request.method == "GET":
        # Create a new Workbook
        wb = Workbook()
        
        # Get the active sheet
        ws = wb.active
        ws.title = "sheet1"
        
        # Define column headers
        columns = ['Id', 'Name', 'Alternate Mode', 'Image']
        
        # Write headers to Excel sheet and set column width
        for col_num, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = col_name
            ws.column_dimensions[cell.column_letter].width = 20  # Adjust column width
        
        # Fetch data from database
        data = Transformer.objects.all().order_by('-id')
        
        # Initialize row counter
        row_num = 2  # Start from row 2 (after headers)
        
        # Loop through data and write to Excel sheet
        for my_row in data:
            # Write data to respective columns
            ws.cell(row=row_num, column=1).value = my_row.id
            ws.cell(row=row_num, column=2).value = my_row.name
            ws.cell(row=row_num, column=3).value = my_row.alternate_mode
            
            # Handle image insertion using openpyxl
            if my_row.image:
                # Example: Assuming my_row.image is a path to the image file
                image_path = my_row.image.path  # Adjust according to your model structure
                
                try:
                    # Open the image using openpyxl
                    img = XLImage(image_path)
                    
                    # Calculate cell dimensions (adjust these as needed)
                    cell_width = 100
                    cell_height = 100
                    
                    # Resize image if necessary to fit within the cell
                    img.width = cell_width
                    img.height = cell_height
                   
                    # Add image to Excel sheet, fix within specific cell (e.g., D2, D3, ...)
                    ws.add_image(img, f'D{row_num}')
                    
                    # Get the cell where the image is inserted
                    cell = ws.cell(row=row_num, column=4)
                    
                    # Set alignment to center the image within the cell
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    img.alignment =cell.alignment
                    
                    # Set row height to accommodate image (adjust as needed)
                    ws.row_dimensions[row_num].height = cell_height
                    
                except Exception as e:
                    print(f"Error processing image: {e}")
            else:
                ws.cell(row=row_num, column=4).value = None

            
            # Increment row counter
            row_num += 1
        
        # Prepare HTTP response for downloading the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="All_Complaints_Report.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        
        return response
